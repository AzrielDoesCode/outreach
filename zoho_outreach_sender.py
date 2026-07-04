

import os
import re
import ssl
import time
import random
import smtplib
from datetime import datetime
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

import openpyxl
from openpyxl.utils import get_column_letter

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# -----------------------------------------------------------------------------
# CONFIG — everything you should check before running
# -----------------------------------------------------------------------------
DRY_RUN = True                     # <-- MUST be flipped to False manually to actually send
BATCH_LIMIT = 10                   # max emails sent in a single run of this script
SEND_DELAY_RANGE = (45, 120)       # seconds between sends, randomized (avoids burst pattern)

XLSX_PATH = "Asian_African_Financial_Institutions_v4_Researched.xlsx"
PREVIEW_DIR = "preview"

SMTP_HOST = os.getenv("ZOHO_SMTP_HOST", "smtppro.zoho.com")   # smtp.zoho.com if not on a paid custom domain
SMTP_PORT = int(os.getenv("ZOHO_SMTP_PORT", "587"))            # 587 = STARTTLS, 465 = SSL
SMTP_USER = os.getenv("ZOHO_SMTP_USER", "")                     # e.g. jsexecutive@xcigence.com
SMTP_APP_PASSWORD = os.getenv("ZOHO_SMTP_APP_PASSWORD", "")    # App-Specific Password, NOT your login password
FROM_NAME = os.getenv("ZOHO_FROM_NAME", "Dhruv Suri, Xcigence")
CC_EMAIL = "yomio@xcigence.com"

# -----------------------------------------------------------------------------
# SIGN-OFF BLOCK (constant across all templates)
# -----------------------------------------------------------------------------
SIGN_OFF_BLOCK = """Dhruv Suri
Enterprise Cybersecurity Resource, Xcigence Corporation

jsexecutive@xcigence.com
Xcigence.com

Patent US11418531B2"""

# -----------------------------------------------------------------------------
# TEMPLATES — subject + body, using {{Column Name}} tokens matching xlsx headers
# -----------------------------------------------------------------------------

TEMPLATE_A_SUBJECT = "{{Organization}} - Standardizing Cyber Risk in {{Region Label}} Portfolios: Xcigence {{Initiative_Short_Name}}"
TEMPLATE_A_BODY = """Dear {{Salutation}},


On behalf of Xcigence, we are glad to share our intent for a strategic
collaboration with {{Organization}} on a critical component that has defined
the sustainability of every organization.

While technology risk remains a critical material threat to enterprise valuation
across {{Region Label}}, institutional investors are still forced to rely on
lagging, static self-disclosures. This information asymmetry is particularly
pressing given that {{Region Label}} data breach costs have climbed
{{Breach Cost YoY %}} YoY to an average of {{Breach Cost Amount}}\u2014
{{Breach Cost Context Clause}}.

To address this, Xcigence Corporation has developed a framework to bring
independent, investor-grade cybersecurity transparency to {{Country_Possessive}}
capital markets.

We have attached our latest brieflink (https://brieflink.com/v/ssd2i): "{{Initiative_Full_Name}}".

Grounded in our patented Cyber Risk Assessment Credit Score (CRACS) methodology
(US Patent No. US11418531B2), this framework translates complex technical risk
into a familiar 300\u2013850 credit-style score. For asset managers, this unlocks:

Portfolio-Wide Visibility: Quantifiable cyber metrics to aggregate and profile
concentration risk across {{Local Market Body}} holdings.

Empirical ESG Integration: An objective, data-driven baseline for the "G"
(Governance) pillar and {{Local Market Body}} sustainability reporting.

Forward-Looking Indicators: The ability to flag deteriorating security postures
before a material breach impacts asset value.


We are convening a 90-minute virtual roundtable with regional
{{Investor Peer Group}} to shape the voluntary {{Initiative_Short_Name}} pilot
framework.

We would welcome your perspective as a founding stakeholder. Would you be open
to a brief introductory call next week to discuss?

Sincerely,

{{SIGN_OFF_BLOCK}}
"""

TEMPLATE_B_SUBJECT = "Briefing: Voluntarily Automating Technology Risk Transparency ({{Initiative_Short_Name}} 2026)"
TEMPLATE_B_BODY = """Dear {{Salutation}},

{{Personal Line (optional)}}

We hope this email finds you well. We are writing on behalf of Xcigence to
explore a strategic collaboration with {{Organization}}. As regulatory
frameworks like {{Regulation 1}}, {{Regulation 2}}, and {{Regulation 3}}
continue to elevate cyber security to a top-tier board priority, the market
still lacks a continuous, standardized baseline to benchmark ecosystem-wide
resilience.

Xcigence Corporation is introducing a collaborative framework outlined in the
attached brieflink (https://brieflink.com/v/ssd2i): "{{Initiative_Full_Name}}".
Our goal is to augment national digital resilience without expanding the
mandatory regulatory burden on local entities.

Utilizing our patented non-intrusive scanning and Risk Assessment methodology,
the platform generates a continuous 300\u2013850 Cyber Risk Assessment Credit Score
(CRACS) that aligns directly with {{Country/Region}}'s core regulatory
mandates:

{{Regulator 1 Short}}: {{Regulator 1 Value Prop}}

{{Regulator 2 Short}}: {{Regulator 2 Value Prop}}


We are launching an 18-month structured pilot program involving 20\u201330
voluntary companies across 5 key sectors, overseen by an independent advisory
panel.

We would highly value inviting {{Self_Reference}} to engage with this
initiative as an official observer to evaluate the data's utility for
evidence-based policy.

May we schedule a brief, 15-minute introductory briefing to walk through the
technical methodology?


Best regards,

{{SIGN_OFF_BLOCK}}
"""

TEMPLATE_C_SUBJECT = "A Cyber Risk Score, Built the Way You Built Credit Risk: Xcigence CRACS"
TEMPLATE_C_BODY = """Dear {{Salutation}},

{{Organization}} has spent years building the infrastructure that lets
{{Country/Region}}'s financial system trust a single number to represent
creditworthiness. Xcigence has built the equivalent for cybersecurity.

Just as a financial credit score condenses years of repayment behavior into a
single, comparable figure, our patented Cyber Risk Assessment Credit Score
(CRACS) \u2014 US Patent No. US11418531B2 \u2014 condenses an organization's entire
technical risk posture into a 300\u2013850 score, generated through continuous,
non-intrusive scanning rather than static self-disclosure.

We believe there is a natural adjacency between what {{Organization}} already
provides to lenders, insurers, and regulators, and what CRACS could add: a
second, independently verifiable dimension of enterprise risk \u2014 digital
resilience \u2014 sitting alongside financial creditworthiness.

A few directions worth exploring together:

Complementary Scoring: CRACS as an additional data layer within your existing
credit reporting products, giving your clients cyber-risk visibility alongside
financial risk.

Shared Methodology Language: Aligning on a consistent, credit-score-style
framework so that "risk" \u2014 financial or cyber \u2014 is communicated to the market
in the same familiar terms.

Joint Market Education: {{Country/Region}}'s lenders and insurers are only
beginning to price cyber risk into their decisions. {{Bureau Adjacency Notes}}


We would welcome a brief, exploratory conversation \u2014 no more than 15 minutes \u2014
to walk through the CRACS methodology and hear how {{Organization}} currently
thinks about non-financial risk data.

Would you be open to a short call in the coming weeks?

Warm regards,

{{SIGN_OFF_BLOCK}}
"""

TEMPLATES = {
    "A": {"subject": TEMPLATE_A_SUBJECT, "body": TEMPLATE_A_BODY},
    "B": {"subject": TEMPLATE_B_SUBJECT, "body": TEMPLATE_B_BODY},
    "C": {"subject": TEMPLATE_C_SUBJECT, "body": TEMPLATE_C_BODY},
}

# -----------------------------------------------------------------------------
# Placeholder rendering
# -----------------------------------------------------------------------------

def derive_fields(row_data):
    """Computes fields that aren't stored directly as xlsx columns."""
    initiative = (row_data.get("Initiative Name") or "").strip()
    country = (row_data.get("Country/Region") or "").strip()
    org = (row_data.get("Organization") or "").strip()

    if initiative == "SECRTI":
        initiative_full = "Singapore Enterprise Cyber Risk Transparency Initiative (SECRTI)"
        initiative_short = "SECRTI"
    else:
        initiative_full = initiative or "Global Investor Cyber Risk Transparency Initiative"
        initiative_short = "GICRTI"

    if country:
        country_possessive = f"{country}'s" if not country.lower().startswith("multiple") else "the region's"
    else:
        country_possessive = "the region's"

    # Self_Reference resolves to exactly one name -- the recipient itself.
    # This is the fix for the unresolved "[MAS/CSA/PDPC]" bug found in the
    # original sent emails.
    self_reference = org

    return {
        "Initiative_Full_Name": initiative_full,
        "Initiative_Short_Name": initiative_short,
        "Country_Possessive": country_possessive,
        "Self_Reference": self_reference,
    }


# Placeholders that are allowed to be blank -- rendering should gracefully
# drop them (and any resulting empty line) rather than error out.
OPTIONAL_PLACEHOLDERS = {"Personal Line (optional)"}


def render(template_str, row_data):
    """Replaces every {{Column Name}} token with the row's value.
    Raises if any REQUIRED token is left unresolved -- we never want to send
    a literal "{{...}}" or a silently-blank required field to a real recipient.
    Fields in OPTIONAL_PLACEHOLDERS are allowed to be empty."""
    merged = dict(row_data)
    merged["SIGN_OFF_BLOCK"] = SIGN_OFF_BLOCK

    def replace(match):
        key = match.group(1).strip()
        value = merged.get(key)
        is_blank = value is None or str(value).strip() == ""

        if key in OPTIONAL_PLACEHOLDERS:
            return "" if is_blank else str(value)

        if is_blank:
            raise ValueError(f"Unresolved or empty placeholder: {{{{{key}}}}}")
        if str(value).strip() == "RESEARCH NEEDED":
            raise ValueError(f"Placeholder still marked RESEARCH NEEDED: {{{{{key}}}}}")
        return str(value)

    rendered = re.sub(r"\{\{([^}]+)\}\}", replace, template_str)

    # Collapse any run of 3+ blank lines left behind by a dropped optional
    # field down to a normal single blank line between paragraphs.
    rendered = re.sub(r"\n{3,}", "\n\n", rendered)

    return rendered


# -----------------------------------------------------------------------------
# xlsx handling
# -----------------------------------------------------------------------------

def load_rows(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    col_idx = {h: i + 1 for i, h in enumerate(headers)}

    # Add tracking columns if they don't exist yet
    for tracking_col in ["Send Status", "Send Timestamp"]:
        if tracking_col not in col_idx:
            new_col = ws.max_column + 1
            ws.cell(row=1, column=new_col, value=tracking_col)
            col_idx[tracking_col] = new_col

    rows = []
    for r in range(2, ws.max_row + 1):
        org = ws.cell(row=r, column=col_idx["Organization"]).value
        if not org:
            continue
        row_data = {h: ws.cell(row=r, column=col_idx[h]).value for h in col_idx}
        row_data["_row_num"] = r
        rows.append(row_data)

    return wb, ws, col_idx, rows


def mark_sent(ws, col_idx, row_num, status):
    ws.cell(row=row_num, column=col_idx["Send Status"], value=status)
    ws.cell(row=row_num, column=col_idx["Send Timestamp"], value=datetime.now().isoformat(timespec="seconds"))


# -----------------------------------------------------------------------------
# Email sending
# -----------------------------------------------------------------------------

def send_email(to_email, subject, body):
    msg = MIMEMultipart()
    msg["From"] = f"{FROM_NAME} <{SMTP_USER}>"
    msg["To"] = to_email
    msg["Cc"] = CC_EMAIL
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain"))

    recipients = [to_email, CC_EMAIL] if CC_EMAIL else [to_email]

    context = ssl.create_default_context()
    if SMTP_PORT == 465:
        with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, context=context) as server:
            server.login(SMTP_USER, SMTP_APP_PASSWORD)
            server.sendmail(SMTP_USER, recipients, msg.as_string())
    else:  # 587, STARTTLS
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            server.starttls(context=context)
            server.login(SMTP_USER, SMTP_APP_PASSWORD)
            server.sendmail(SMTP_USER, recipients, msg.as_string())


# -----------------------------------------------------------------------------
# Main
# -----------------------------------------------------------------------------

def main():
    if not os.path.exists(PREVIEW_DIR):
        os.makedirs(PREVIEW_DIR)

    wb, ws, col_idx, rows = load_rows(XLSX_PATH)

    sent_this_run = 0
    skipped_template = 0
    skipped_already_sent = 0
    errors = []

    for row_data in rows:
        row_num = row_data["_row_num"]
        org = row_data.get("Organization", "UNKNOWN")
        label = (row_data.get("Template Label") or "").strip()
        email_addr = row_data.get("Email")
        already_sent = row_data.get("Send Status") == "Sent"

        if label not in TEMPLATES:
            skipped_template += 1
            continue  # skips D and UNASSIGNED silently -- by design

        if already_sent:
            skipped_already_sent += 1
            continue

        if not email_addr:
            errors.append((org, "No email address on file"))
            continue

        derived = derive_fields(row_data)
        merged = {**row_data, **derived}

        try:
            subject = render(TEMPLATES[label]["subject"], merged)
            body = render(TEMPLATES[label]["body"], merged)
        except ValueError as e:
            errors.append((org, str(e)))
            mark_sent(ws, col_idx, row_num, f"FAILED: {e}")
            continue

        # Always write the preview file, dry run or not
        safe_name = re.sub(r"[^\w\-]+", "_", org)[:60]
        preview_path = os.path.join(PREVIEW_DIR, f"{safe_name}.txt")
        with open(preview_path, "w", encoding="utf-8") as f:
            f.write(f"TO: {email_addr}\nCC: {CC_EMAIL}\nSUBJECT: {subject}\n\n{body}")

        if DRY_RUN:
            print(f"[DRY RUN] Would send to {org} <{email_addr}> — preview saved to {preview_path}")
            continue

        if sent_this_run >= BATCH_LIMIT:
            print(f"[STOP] Batch limit of {BATCH_LIMIT} reached for this run. "
                  f"Run the script again later to continue with the next batch.")
            break

        try:
            send_email(email_addr, subject, body)
            mark_sent(ws, col_idx, row_num, "Sent")
            sent_this_run += 1
            print(f"[SENT] {org} <{email_addr}>")
        except Exception as e:
            mark_sent(ws, col_idx, row_num, f"FAILED: {e}")
            errors.append((org, str(e)))
            print(f"[FAIL] {org} <{email_addr}> — {e}")
            continue

        delay = random.uniform(*SEND_DELAY_RANGE)
        print(f"    ...waiting {delay:.0f}s before next send")
        time.sleep(delay)

    wb.save(XLSX_PATH)

    print("\n" + "=" * 60)
    print(f"DRY_RUN = {DRY_RUN}")
    print(f"Sent this run: {sent_this_run}")
    print(f"Skipped (Template D / unassigned): {skipped_template}")
    print(f"Skipped (already sent previously): {skipped_already_sent}")
    print(f"Errors: {len(errors)}")
    for org, err in errors:
        print(f"  - {org}: {err}")
    print("=" * 60)


if __name__ == "__main__":
    main()
